# 李文胜改
import time
import os
import random
import pygame
from openpyxl import *
from time import sleep
#背景音乐
pygame.mixer.init()
pygame.mixer.music.load(r'川井憲次 - 孤独な巡礼.mp3')
pygame.mixer.music.play(-1, 0, 0)

def Sound(S):  # 答题声音
    
    pygame.init()
    soundObj = pygame.mixer.Sound(S)
    soundObj.play()


def practice():  # 练习模式函数
    Sound(S24)
    print("你最好选择适合你的模式哦，才...才没有为你担心呢！\n汉译英模式A\t英译汉模式B")

    P = input("")
    while True:

        Sound(S29)


        O = input("要练多少次才好呢，这次勉强让你来决定吧：")

        sleep(2)

        Sound(S19)
        print('美~全力以赴吧！你可不准说"No 哦~"')
        sleep(1)
        if O.isdigit():
            return P.upper(), int(O)
            break
        else:
            Sound(S21)
            print(
                """
                笨蛋笨蛋，你输入的不规范，给我重新输啦！！！  ಥ◡ಥ
                """
            )
            continue


def chinese_english_practice(O):  # 模式函数
    n = input("按回车就会开始喽！")
    while n == '':
        wrong = []  # 创建一个装错题的列表
        for i in range(0, O):
            L = list(words.items())
            x = random.randint(0, len(L) - 1)  # !
            key = (L[x][0])
            right_key = L[x][1]
            print("\n" + key)
            user_key = input("")
            if user_key == right_key:
                Sound(S28)
                print("你这个笨蛋居然答对了？？？")
            else:
                wrong.append((L[x]))
                Sound(S22)
                print("！回答错啦，准备接受惩罚吧！\n勉强告诉你正确答案是：{}".format(right_key))
        n = '1'


def english_chinese_pratice(O):
    n = input("按回车就会开始喽！")  # 英译汉模式函数
    while n == '':
        wrong = []
        for i in range(0, O):
            L = list(words.items())
            x = random.randint(0, len(L) - 1)
            value = (L[x][1])  # 从库列表中取一个数
            right_value = L[x][0]
            print("\n" + value)
            user_value = input("")
            if user_value == right_value:
                Sound(S35)
                print("居...居然对了？算你运气好")
            else:
                wrong.append((L[x]))
                Sound(S22)

                print("笨蛋，这都不会，告诉你正确答案是：{}，给我记好喽！".format(right_value))
        n = '1'


def exam():  # 考试系统函数
    Sound(S29)
    print("选择一个练习模式吧~\n汉译英A\t英译汉B\t混合模式C")
    P = input("")
    Sound(S18)
    print("想要练习几次咧？才没有为你担心呢！：")
    O = input("")
    return P.upper(), int(O)


def chinese_english_exam(O):  # 英译汉模式函数
    n = input("按回车开始吧")
    while n == '':
        z = 0  # 记录正确题数
        wrong = list()
        for i in range(0, O):
            L = list(words.items())
            x = random.randint(0, len(L) - 1)
            key = (L[x][0])  # 选中题目
            right_key = L[x][1]  # 设定答案
            print("\n" + key)
            user_key = input("")
            if user_key == right_key:
                z = z + 1
            else:
                wrong.append((L[x]))
        if O - z == 0:
            Sound(S11)
            print("考试结束喽，一共完成{}题。我觉绝对没有为你高兴哦。\n正确率100%".format(O))
        else:
            print(
                '考试结束，一共完成{}题。\n正确{}题，错误{}题\n正确率为:{:.2%}'.format(O, z, O - z, z / O))
            Sound(S2)
            print("错误单词如下{}，错太多啦！笨蛋笨蛋！大笨蛋！再也不理你了！".format(wrong))
            sleep(3)
        n = '1'


def english_chinese_exam(O):  # 汉译英模式函数
    n = input("按回车就会开启我们的旅程！")
    while n == '':
        wrong = list()
        z = 0  # 记录正确题数
        for i in range(0, O):
            L = list(words.items())
            x = random.randint(0, len(L) - 1)
            value = (L[x][1])  # 从库列表中取一个数
            right_value = L[x][0]
            print("\n" + value)
            user_value = input("")
            if user_value == right_value:
                z = z + 1
            else:
                wrong.append((L[x]))
        if O - z == 0:
            print("考试结束，一共完成{}题。\n正确率100%".format(O))
        else:
            print(
                '考试结束，一共完成{}题。\n正确{}题，错误{}题\n您的正确率为:{:.2%}'.format(O, z, O - z, z / O))
            print("错误单词如下{}".format(wrong))
            Sound(S7)
            print('笨蛋，背单词不能着急哦，一定要背熟呢！')
        n = '1'


def mixed_exam(O):  # 混合模式函数
    n = input("按回车开始旅程吧！")
    while n == '':
        wrong = list()
        z = 0  # 记录正确题数
        for i in range(0, O):
            L = list(words.items())
            x = random.randint(0, len(L) - 1)
            y = random.randint(0, 1)
            if y == 1:  # see_chinese
                value = (L[x][1])  # 从库列表中取一个数
                right_value = L[x][0]
                print("\n" + value)
                user_value = input("")
                if user_value == right_value:
                    z = z + 1
                else:
                    wrong.append((L[x]))
            elif y == 0:  # see_english
                value = (L[x][0])  # 从库列表中取一个数
                right_value = L[x][1]
                print("\n" + value)
                user_value = input("")
                if user_value == right_value:
                    z = z + 1
                else:
                    wrong.append((L[0]))
        if O - z == 0:
            print("本次考试结束，一共完成{}题。\n正确率100%".format(O))
        else:

            print(
                '本次考试结束，一共完成{}题。\n其中正确{}题，错误{}题\n本次您的正确率为:{:.2%} 你这家伙错这么多，找块豆腐撞死得了！'.format(O, z, O - z, z / O))
            Sound(S8)
            print("错误单词如下{}".format(wrong))
            sleep(3)
        n = '1'


def main():
    print(                '时光回溯.ing（倒计时5秒）  原神，启动！！！'                      )
    print(5)
    sleep(1)
    print(4)
    sleep(1)
    print(3)
    sleep(1)
    print(2)
    sleep(1)
    print(1)
    sleep(1)
    Sound(S9)
    print(
        """
                                         全力以赴吧！！！我会看着你的！
                                                  主菜单 
                          练习按A          考试按B          题库按C          退出按D
        """
    )
    return input()




# 从词库获取列表函数
te1 = load_workbook(u"data.xlsx")
sheetnames1 = te1.sheetnames
t1 = te1['Sheet1']
words = {}
for i in range(1, t1.max_row + 1):
    english = t1[f'A{i}'].value  # 英文
    chinese = t1[f'B{i}'].value  # 中文
    words[english] = chinese
S1 = r'傲娇\01.被偷的是我的…真是的，不说了啦笨蛋!!你让人家怎么说嘛!!.wav'
S2 = r'傲娇\02.变态 变态 变态 变态 变态 你这个大变态!!我再也不理你了….wav'
S3 = r'03.别把拉面挑这么高，快点吃！ 什么？帮你吹凉一点？去死！.wav'
S4 = r'傲娇\04.别不理我嘛…我道歉还不行吗….wav'
S5 = r'傲娇\05.别离我这么近…丢死人了….wav'
S7 = r'傲娇\07.吵死了吵死了吵死了! 真是吵死了!.wav'
S8 = r'傲娇\08.给我等一下! 你这个木头人，为什么…为什么你就是注意不到嘛!.wav'
S9 = r'傲娇\09.尽管放马过来吧! 让我打你个落花流水!.wav'
S11 = r'傲娇\11.就算你送我琉璃色的发带，我也绝对不会开心的哦!.wav'
S12 = r'傲娇\12.哭…我的胸部….wav'
S13 = r'傲娇\13.留在我身边! 一辈子! 没异议吧!.wav'
S16 = r'傲娇\16.你、你只要看着我一个人就够了!.wav'
S18 = r'傲娇\18.你可别胡思乱想，我有不是特地为你做的.wav'
S21 = r"傲娇\21.你这个笨蛋! 根本不知道我的想法….wav"
S19 = r"傲娇\19.你可不许对我说“No”哦.wav"
S22 = r'傲娇\22.你这家伙，找块豆腐一头撞死算了!.wav'
S24 = r'傲娇\24.你至少和我联系一下啊…人家好担心的说….wav'
S28 = r'傲娇\28.什么利用价值都没有，你这种废柴.wav'
S29 = r'傲娇\29.是我选的你，你可别以为是你给了我机会哦!.wav'
S32 = r'傲娇\32.我、我才没有为你担心!.wav'
S35 = r'傲娇\35.我才没有吃醋!都给你说了没有了嘛!.wav'
S41 = r'傲娇\41.真是讨厌讨厌讨厌!!你要想做那种事情我不理你了不理你了!.wav'


# todo 2
print('                                                记words系统                                               ')



while True:


    t = main()


    if t.upper() == 'A':
        a, b = practice()
        if a == "A":
            english_chinese_pratice(b)
        elif a == "B":
            chinese_english_practice(b)
        else:
            Sound(S21)
            print(
                """
                笨蛋！你会不会敲键盘啊！输的不规范重新输啦！  ಥ◡ಥ

                """
            )
            continue
        continue
    elif t.upper() == 'B':
        a, b = exam()
        if a == "A":
            english_chinese_exam(b)
        elif a == "B":
            chinese_english_exam(b)
        elif a == "C":
            mixed_exam(b)

        else:
            Sound(S21)
            print(
                """
                笨蛋！你会不会敲键盘啊！输的不规范重新输啦！  ಥ◡ಥ  ಥ◡ಥ
                """
            )
            continue
        continue
        sleep(3)
    elif t.upper() == 'C':
        os.system('data.xlsx')
        continue
    elif t.upper() == 'D':
        Sound(13)
        print('怎么...怎么就要离开了？留在我身边! 一辈子! 没异议吧!    ')
        sleep(5)
        break
    else:
        Sound(S21)
        print(
            """
            笨蛋！你会不会敲键盘啊！输的不规范重新输啦！  ಥ◡ಥ 
            """
        )
